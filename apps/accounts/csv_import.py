"""Shared bulk-import plumbing for the setup masters (spec §1 onboarding).

Every master import works the same way: GET the endpoint for a fill-in CSV
template (opens in Excel), POST it back as `file` (CSV or XLSX) and get a
per-row report — created / skipped-existing / errors — never an all-or-nothing
failure halfway through onboarding.
"""
import csv
import io

from django.http import HttpResponse


def template_response(filename, columns, examples):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    for row in examples:
        w.writerow(row)
    resp = HttpResponse(buf.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def export_response(filename_base, header, rows, fmt="xlsx"):
    """CSV or XLSX download of `rows` under `header`. Mirrors
    ReportExportView's two-format branch, kept local to the setup
    masters so a report-view change never touches this."""
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(header)
        for row in rows:
            w.writerow(row)
        resp = HttpResponse(buf.getvalue(), content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        return resp
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append([str(c) for c in row])
    out = io.BytesIO()
    wb.save(out)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
    return resp


def parse_upload(request):
    """Rows from an uploaded CSV/XLSX `file` as (lineno, {column: value}),
    keyed by the header row. Raises ValueError with a user-facing message."""
    f = request.FILES.get("file")
    if not f:
        raise ValueError("attach a CSV or XLSX file as 'file'")
    try:
        if f.name.lower().endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            ws = load_workbook(io.BytesIO(f.read()), read_only=True).active
            data = [[("" if c is None else str(c)) for c in row]
                    for row in ws.iter_rows(values_only=True)]
        else:
            text = f.read().decode("utf-8-sig", errors="replace")
            data = list(csv.reader(io.StringIO(text)))
    except Exception:
        raise ValueError("could not read the file — use the downloaded template")
    if not data or len(data) < 2:
        raise ValueError("the file has no data rows")
    header = [h.strip().lower().replace(" ", "_") for h in data[0]]
    return [
        (lineno, {header[i]: (raw[i] or "").strip()
                  for i in range(min(len(header), len(raw)))})
        for lineno, raw in enumerate(data[1:], start=2)
    ]
