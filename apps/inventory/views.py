import csv
import io
from datetime import timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.accounts.permissions import ModuleViewSetMixin

from .models import Ingredient, IngredientCategory, StockMovement, Uom, apply_movement
from .serializers import (
    IngredientCategorySerializer,
    IngredientSerializer,
    StockMovementSerializer,
    UomSerializer,
)


class UomViewSet(ModuleViewSetMixin, viewsets.ModelViewSet):
    """Units of Measurement master (spec §6)."""

    module = "inventory"
    queryset = Uom.objects.all()
    serializer_class = UomSerializer

    def destroy(self, request, *args, **kwargs):
        uom = self.get_object()
        used = Ingredient.objects.filter(unit=uom.code).count()
        if used:
            return Response({"detail": f"'{uom.code}' is used by {used} material(s)"}, status=400)
        return super().destroy(request, *args, **kwargs)


class IngredientCategoryViewSet(ModuleViewSetMixin, viewsets.ModelViewSet):
    """Raw-material categories master (spec §6)."""

    module = "inventory"
    queryset = IngredientCategory.objects.all()
    serializer_class = IngredientCategorySerializer

    def destroy(self, request, *args, **kwargs):
        cat = self.get_object()
        used = Ingredient.objects.filter(category=cat.name).count()
        if used:
            return Response({"detail": f"'{cat.name}' is used by {used} material(s)"}, status=400)
        return super().destroy(request, *args, **kwargs)


class IngredientViewSet(ModuleViewSetMixin, viewsets.ModelViewSet):
    module = "inventory"
    queryset = Ingredient.objects.all()
    serializer_class = IngredientSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.query_params.get("below_par") == "1":
            # below_par is a property, so resolve to ids the DB can filter on
            ids = [i.id for i in qs if i.below_par]
            return qs.filter(id__in=ids)
        return qs

    @action(detail=True, methods=["post"])
    def adjust(self, request, pk=None):
        """Manual stock adjustment (spec §4: Stock Adjustment)."""
        ing = self.get_object()
        apply_movement(ing, StockMovement.ADJUSTMENT, request.data.get("qty", 0),
                       reason=request.data.get("reason", ""), user=request.user)
        return Response(IngredientSerializer(ing).data)

    @action(detail=True, methods=["post"])
    def waste(self, request, pk=None):
        """Wastage entry — always stock-out, reason required (spec §4/§6)."""
        ing = self.get_object()
        qty = abs(Decimal(str(request.data.get("qty", 0))))
        reason = request.data.get("reason", "").strip()
        if qty <= 0:
            return Response({"detail": "quantity must be positive"}, status=400)
        if not reason:
            return Response({"detail": "a reason is required for wastage"}, status=400)
        kind = (StockMovement.EXPIRY if request.data.get("expired")
                else StockMovement.WASTAGE)
        apply_movement(ing, kind, -qty, reason=reason, user=request.user)
        return Response(IngredientSerializer(ing).data)

    @action(detail=True, methods=["post"])
    def transfer(self, request, pk=None):
        """Stock transfer (spec §4/§6): move stock out to (or in from) another
        location/outlet. direction=out issues stock, direction=in receives it;
        the movement's reference records the counterparty location."""
        ing = self.get_object()
        try:
            qty = abs(Decimal(str(request.data.get("qty", 0))))
        except Exception:
            return Response({"detail": "invalid quantity"}, status=400)
        location = (request.data.get("location") or "").strip()
        direction = request.data.get("direction", "out")
        if qty <= 0:
            return Response({"detail": "quantity must be positive"}, status=400)
        if not location:
            return Response({"detail": "the other location is required"}, status=400)
        if direction not in ("out", "in"):
            return Response({"detail": "direction must be out or in"}, status=400)
        if direction == "out" and qty > (ing.current_stock or Decimal("0")):
            return Response({"detail": "not enough stock to transfer out"}, status=400)
        signed = -qty if direction == "out" else qty
        apply_movement(ing, StockMovement.TRANSFER, signed,
                       reason=("to " if direction == "out" else "from ") + location,
                       source=f"transfer:{location}", user=request.user)
        return Response(IngredientSerializer(ing).data)

    @action(detail=True, methods=["post"])
    def count(self, request, pk=None):
        """Physical stock count: book the counted quantity, the difference is
        posted as a 'count' movement so the ledger explains the correction (§4/§6)."""
        ing = self.get_object()
        counted = Decimal(str(request.data.get("counted", 0)))
        if counted < 0:
            return Response({"detail": "counted quantity can't be negative"}, status=400)
        diff = counted - (ing.current_stock or Decimal("0"))
        if diff != 0:
            apply_movement(ing, StockMovement.COUNT, diff,
                           reason=f"physical count: booked {counted}", user=request.user)
        return Response(IngredientSerializer(ing).data)

    @action(detail=False, methods=["get", "post"], url_path="import")
    def import_materials(self, request):
        """Bulk onboarding (spec §1): upload a CSV/XLSX of raw materials.

        GET returns the fill-in template. POST with a `file` creates every row,
        auto-creating unknown categories/units (onboarding shouldn't fight the
        masters), books opening stock through the ledger, and reports per-row
        errors instead of failing the whole file.
        """
        columns = ["name", "category", "unit", "opening_stock", "min_stock_level",
                   "reorder_level", "purchase_rate", "storage_location", "expiry_date"]
        if request.method == "GET":
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(columns)
            w.writerow(["Basmati Rice", "Other Consumables", "kg", "25", "5", "10",
                        "90", "Dry store", ""])
            w.writerow(["Milk", "Dairy", "l", "20", "5", "10", "60", "Cold room",
                        "2026-07-15"])
            resp = HttpResponse(buf.getvalue(), content_type="text/csv")
            resp["Content-Disposition"] = 'attachment; filename="raw-materials-template.csv"'
            return resp

        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "attach a CSV or XLSX file as 'file'"}, status=400)
        # Parse into a list of dicts keyed by the template's column names.
        try:
            if f.name.lower().endswith((".xlsx", ".xlsm")):
                from openpyxl import load_workbook
                ws = load_workbook(io.BytesIO(f.read()), read_only=True).active
                data = [[("" if c is None else str(c)) for c in row]
                        for row in ws.iter_rows(values_only=True)]
            else:
                text = f.read().decode("utf-8-sig", errors="replace")
                data = list(csv.reader(io.StringIO(text)))
        except Exception:
            return Response({"detail": "could not read the file — use the template"}, status=400)
        if not data or len(data) < 2:
            return Response({"detail": "the file has no data rows"}, status=400)
        header = [h.strip().lower().replace(" ", "_") for h in data[0]]

        created, skipped, errors = [], [], []
        for lineno, raw in enumerate(data[1:], start=2):
            row = {header[i]: (raw[i] or "").strip() for i in range(min(len(header), len(raw)))}
            name = row.get("name", "")
            if not name:
                continue  # blank line
            if Ingredient.objects.filter(name__iexact=name).exists():
                skipped.append(name)
                continue
            try:
                unit = (row.get("unit") or "kg").lower()
                Uom.objects.get_or_create(code=unit, defaults={"name": unit})
                category = row.get("category", "")
                if category:
                    IngredientCategory.objects.get_or_create(name=category)
                expiry = row.get("expiry_date") or None
                ing = Ingredient.objects.create(
                    name=name, unit=unit, category=category,
                    min_stock_level=Decimal(row.get("min_stock_level") or "0"),
                    reorder_level=Decimal(row.get("reorder_level") or "0"),
                    unit_cost=Decimal(row.get("purchase_rate") or row.get("unit_cost") or "0"),
                    storage_location=row.get("storage_location", ""),
                    expiry_date=expiry[:10] if expiry else None,
                )
                opening = Decimal(row.get("opening_stock") or "0")
                if opening > 0:
                    apply_movement(ing, StockMovement.ADJUSTMENT, opening,
                                   reason="opening stock (import)", user=request.user)
                created.append(name)
            except Exception as e:
                errors.append({"row": lineno, "name": name, "reason": str(e)[:120]})
        return Response({"created": len(created), "skipped_existing": skipped,
                         "errors": errors})

    @action(detail=False, methods=["get"])
    def low_stock(self, request):
        low = [i for i in Ingredient.objects.all() if i.below_par]
        return Response({"count": len(low), "items": IngredientSerializer(low, many=True).data})

    @action(detail=False, methods=["get"])
    def expiring(self, request):
        """Materials expiring within ?days= (default 7) or already expired (§6)."""
        days = int(request.query_params.get("days", 7))
        horizon = timezone.localdate() + timedelta(days=days)
        items = Ingredient.objects.filter(expiry_date__isnull=False,
                                          expiry_date__lte=horizon).order_by("expiry_date")
        return Response(IngredientSerializer(items, many=True).data)

    @action(detail=False, methods=["get"])
    def movements(self, request):
        """Inventory movements / consumption register (spec §4/§6).
        Filters: ?kind=consumption&ingredient=<id>&days=30 or an explicit
        ?from=YYYY-MM-DD&to=YYYY-MM-DD range. ?fmt=csv downloads the register."""
        qs = StockMovement.objects.select_related("ingredient")
        kind = request.query_params.get("kind")
        if kind:
            qs = qs.filter(kind=kind)
        ingredient = request.query_params.get("ingredient")
        if ingredient:
            qs = qs.filter(ingredient_id=ingredient)
        date_from = parse_date(request.query_params.get("from") or "")
        date_to = parse_date(request.query_params.get("to") or "")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if not (date_from or date_to):
            days = int(request.query_params.get("days", 30))
            qs = qs.filter(created_at__gte=timezone.now() - timedelta(days=days))

        if request.query_params.get("fmt") == "csv":
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(["Date", "Material", "Type", "Qty", "Unit", "Balance",
                        "Reference", "Reason", "By"])
            for m in qs[:5000]:
                w.writerow([m.created_at.strftime("%Y-%m-%d %H:%M"), m.ingredient.name,
                            m.get_kind_display(), m.qty, m.ingredient.unit, m.balance,
                            m.source, m.reason, m.created_by])
            resp = HttpResponse(buf.getvalue(), content_type="text/csv")
            resp["Content-Disposition"] = 'attachment; filename="consumption-register.csv"'
            return resp
        return Response(StockMovementSerializer(qs[:500], many=True).data)

    @action(detail=False, methods=["get"])
    def consumption_report(self, request):
        """Raw-material consumption + purchase-vs-consumption over ?days= (spec §7)."""
        from apps.accounts.constants import ROLE_CHEF
        show_cost = getattr(request.user, "role", "") != ROLE_CHEF
        days = int(request.query_params.get("days", 30))
        since = timezone.now() - timedelta(days=days)
        rows = []
        for ing in Ingredient.objects.all():
            sums = (ing.movements.filter(created_at__gte=since)
                    .values("kind").annotate(total=Sum("qty")))
            by_kind = {r["kind"]: r["total"] or Decimal("0") for r in sums}
            consumed = -(by_kind.get(StockMovement.CONSUMPTION, Decimal("0")))
            wasted = -(by_kind.get(StockMovement.WASTAGE, Decimal("0"))
                       + by_kind.get(StockMovement.EXPIRY, Decimal("0")))
            purchased = by_kind.get(StockMovement.RECEIPT, Decimal("0"))
            if not (consumed or wasted or purchased):
                continue
            cost = consumed * (ing.unit_cost or Decimal("0"))
            rows.append({
                "ingredient": ing.name, "code": ing.code, "unit": ing.unit,
                "consumed": consumed, "wasted": wasted, "purchased": purchased,
                # Chef's reason to be here is tracking usage/wastage, not the
                # cost — same rule as the ingredient rate (see IngredientSerializer).
                "consumption_cost": (cost if show_cost else None),
                "in_stock": ing.current_stock,
                "_sort_cost": cost,
            })
        rows.sort(key=lambda r: -r["_sort_cost"])
        for r in rows:
            del r["_sort_cost"]
        return Response({"days": days, "rows": rows})
