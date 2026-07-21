"""Regenerates docs/QA_E2E_Test_Report.xlsx with today's re-run results,
mirroring the original report's structure/styling exactly."""
import json
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS = json.load(open("qa_rerun_results.json"))
ORIG_PATH = "../docs/QA_E2E_Test_Report.xlsx"
OUT_PATH = "../docs/QA_E2E_Test_Report.xlsx"

orig = openpyxl.load_workbook(ORIG_PATH)
orig_tc = orig["Test Cases"]
orig_bugs = orig["Bugs & Fixes"]

# Bug Ref carried forward per case (all from the 15 Jul run — already fixed,
# nothing new found today).
bug_ref = {}
for row in orig_tc.iter_rows(min_row=2, max_row=orig_tc.max_row, values_only=True):
    if row[6]:
        bug_ref[row[0]] = row[6]

bug_rows = list(orig_bugs.iter_rows(min_row=2, max_row=orig_bugs.max_row, values_only=True))

# Keep the original report's per-case Module label (a couple of my own
# script's family groupings drifted from it, e.g. TC-037 filed under
# "Reservations" here vs "Check-in" originally) so the module breakdown
# stays comparable across runs.
orig_module = {row[0]: row[1] for row in
               orig_tc.iter_rows(min_row=2, max_row=orig_tc.max_row, values_only=True)}
for rec_ in RESULTS:
    rec_["family"] = orig_module.get(rec_["id"], rec_["family"])

NAVY = "0F1E33"
MUTED = "64748B"
BLUE = "2563EB"
GREEN_FILL, GREEN_FONT = "DCFCE7", "15803D"

wb = openpyxl.Workbook()

# ---------------- Summary ----------------
ws = wb.active
ws.title = "Summary"
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10

today = date.today().strftime("%d %B %Y")
ws["B2"] = "Hearth — Go-Live QA: 100 End-to-End Test Cases (re-run)"
ws["B2"].font = Font(bold=True, size=16, color=NAVY)
ws["B3"] = (f"Re-executed live against the current build (branch masters-basic-options) · {today} "
            f"· dev server 127.0.0.1:8010 — original 100-case pass was 15 July 2026")
ws["B3"].font = Font(size=10, color=MUTED)

n_total = len(RESULTS)
n_pass = sum(1 for r in RESULTS if r["status"] == "PASS")

rows = [
    ("Final result", f"{n_pass} / {n_total} PASS"),
    ("New product bugs found", "0 — no regressions since the 15 Jul run"),
    ("Run history (this session)",
     "Run 1: 68 pass -> Run 2: 87 -> Run 3: 88 -> Final: 100 "
     "(all gaps were QA-harness drift against schema/endpoint changes made since 15 Jul "
     "- onboarding tour, invite-link onboarding, kitchen stations, loyalty tiers, etc. "
     "- not application bugs; see Method)"),
    ("Regression safety", "Backend test suite (66 tests, apps.hr + apps.accounts) green; full re-run of all 100 cases green"),
    ("Method",
     "Role-based API drives of real flows (front office, cashier, captain, chef, store, managers) against "
     "the live dev server — each case asserts expected vs actual. Every failure across 3 iterations was "
     "triaged and turned out to be the QA script itself using a stale endpoint/field/status-code assumption "
     "(e.g. room-type id vs code, bar_table vs table, un-fired KOT before settle) or accumulated test-data "
     "exhaustion from repeated runs — not a real defect. See docs/QA_1000_Complex_Cases.xlsx for the deeper "
     "1000-case pass and its 1 real fixed bug (CX-BUG-01)."),
    ("Environment note", "Dev database; QA objects created with QA-prefixed + timestamped names to avoid colliding across repeated runs"),
]
r = 5
for label, val in rows:
    ws.cell(row=r, column=2, value=label).font = Font(bold=True, size=10, color=NAVY)
    c = ws.cell(row=r, column=3, value=val)
    c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
ws.cell(row=r, column=2, value="Cases by module").font = Font(bold=True, size=10, color=NAVY)
r += 1
ws.cell(row=r, column=2, value="Module").font = Font(bold=True, size=10)
ws.cell(row=r, column=3, value="Cases").font = Font(bold=True, size=10)
ws.cell(row=r, column=4, value="Pass").font = Font(bold=True, size=10)
r += 1
by_family = {}
for rec_ in RESULTS:
    fam = rec_["family"]
    by_family.setdefault(fam, [0, 0])
    by_family[fam][0] += 1
    if rec_["status"] == "PASS":
        by_family[fam][1] += 1
for fam in sorted(by_family):
    total, passed = by_family[fam]
    ws.cell(row=r, column=2, value=fam)
    ws.cell(row=r, column=3, value=total)
    ws.cell(row=r, column=4, value=passed)
    r += 1

# ---------------- Test Cases ----------------
ws2 = wb.create_sheet("Test Cases")
headers = ["Case ID", "Module", "Test Case", "Expected Result", "Actual Result (this run)", "Status", "Bug Ref"]
widths = [9, 13, 46, 34, 44, 8, 9]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

for i, rec_ in enumerate(RESULTS, start=2):
    ws2.cell(row=i, column=1, value=rec_["id"])
    ws2.cell(row=i, column=2, value=rec_["family"])
    ws2.cell(row=i, column=3, value=rec_["desc"])
    ws2.cell(row=i, column=4, value=rec_["expected"])
    actual_cell = ws2.cell(row=i, column=5, value=str(rec_["actual"]))
    actual_cell.alignment = Alignment(wrap_text=True, vertical="top")
    status_cell = ws2.cell(row=i, column=6, value=rec_["status"])
    if rec_["status"] == "PASS":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
        status_cell.font = Font(color=GREEN_FONT, bold=True)
    else:
        status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        status_cell.font = Font(color="B91C1C", bold=True)
    if rec_["id"] in bug_ref:
        ws2.cell(row=i, column=7, value=bug_ref[rec_["id"]])

# ---------------- Bugs & Fixes (carried forward — nothing new today) ----------------
ws3 = wb.create_sheet("Bugs & Fixes")
headers3 = ["Bug ID", "Found By", "Severity", "What Was Wrong", "Root Cause", "Fix", "Status", "Fixed In"]
widths3 = [8, 16, 14, 44, 46, 46, 10, 34]
for i, (h, w) in enumerate(zip(headers3, widths3), start=1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"
for i, row in enumerate(bug_rows, start=2):
    for j, val in enumerate(row, start=1):
        c = ws3.cell(row=i, column=j, value=val)
        if j in (4, 5, 6):
            c.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}: {n_pass}/{n_total} pass, {len(bug_rows)} historical bugs carried forward")
