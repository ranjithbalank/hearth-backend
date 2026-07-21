"""Regenerates docs/QA_1000_Complex_Cases.xlsx with today's re-run results,
mirroring the original report's structure/styling. Unlike the original run
(~165 accidental duplicate role x endpoint rows), this run's RBAC-read family
was generated from a de-duplicated URL-introspected endpoint catalogue —
every case appears exactly once."""
import json
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS = json.load(open("qa1000_results.json"))
OUT_PATH = "../docs/QA_1000_Complex_Cases.xlsx"

NAVY = "0F1E33"
MUTED = "64748B"
BLUE = "2563EB"
GREEN_FILL, GREEN_FONT = "DCFCE7", "15803D"

FAMILY_DESC = {
    "RBAC-read": "Role x endpoint read matrix — every role against every module-gated GET endpoint "
                 "(discovered by introspecting the live URLconf), expected 200/403 derived live from "
                 "apps.accounts.rbac.can_access() — the same DB-aware check the API itself enforces",
    "RBAC-write": "Config-write denial — every role lacking the owning module blocked from settings/"
                  "masters/roles/branches/tax/inventory writes",
    "Identity": "Real login (password, not force_authenticate) + /me role claim + wrong-password "
                "rejection, per role",
    "Reports": "Report endpoints scoped per ROLE_REPORT_ACCESS — full-access roles see every report, "
               "everyone else only their slice",
    "Masters": "Masters guards — seeded data, CRUD, builtin-tender rename/delete protection, in-use "
               "delete block, cross-role write denial",
    "Currency": "Property currency round-trip across all 12 supported currencies",
    "Entitlement": "CX-RBAC-02 regression (non-settings role denied entitlement write) + edition "
                   "gating on/off (banquets blocked/restored)",
    "Audit": "Audit-trail presence for key actions (masters, entitlement) + immutability (no delete route)",
    "Tender": "ROLE_TENDERS ground truth exercised live — role x tender settle allow/deny, including "
              "the Captain till/reconciliation counter-only carve-out and Gateway's payment-token path",
    "Numbering": "Document-numbering integrity — invoice/PO/BEO sequences consecutive, unique, "
                 "correctly formatted",
    "State": "Illegal state transitions — double check-in, self-approve indent, store-keeper "
             "approve-block, receive-before-approve, idempotent re-checkout (no double charge)",
    "KYC": "KYC/registration evidence — flags not blobs, audited retrieval",
    "Lifecycle": "Employee/leave/banquet/housekeeping/work-order/material-request/PO-GRN lifecycles, "
                 "including two-level leave approval and requester != approver != issuer segregation",
    "Money": "GST line-total consistency, discount-cap enforcement, void-after-KOT authorisation guard",
    "DPDP": "Guest data export + erasure anonymisation + role gating",
    "Validation": "Input-validation negatives — malformed/duplicate/blank/weak payloads rejected with 400",
}

FAMILY_ORDER = ["RBAC-read", "RBAC-write", "Identity", "Reports", "Validation", "Currency", "Masters",
                "Lifecycle", "Tender", "Numbering", "State", "Audit", "Money", "KYC", "DPDP", "Entitlement"]

wb = openpyxl.Workbook()

# ---------------- Summary ----------------
ws = wb.active
ws.title = "Summary"
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 90
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 10

today = date.today().strftime("%d %B %Y")
ws["B2"] = "Hearth — 1000+ Complex End-to-End Cases (re-run)"
ws["B2"].font = Font(bold=True, size=16, color=NAVY)
ws["B3"] = (f"Re-executed in-process against the current build (branch masters-basic-options) · {today} "
            f"— original 1195-case pass was 15 July 2026")
ws["B3"].font = Font(size=10, color=MUTED)

n_total = len(RESULTS)
n_pass = sum(1 for r in RESULTS if r["status"] == "PASS")

rows = [
    ("Final result", f"{n_pass} / {n_total} PASS"),
    ("New bug found & fixed",
     "1 (High/security) — CX-RBAC-02: PATCH /auth/entitlements/ was gated only by IsAuthenticated "
     "despite its own docstring claiming server-side MD/GM enforcement — any logged-in role "
     "(Housekeeping, a cashier, a captain) could flip HMS/restaurant/banquets/RMS entitlements for the "
     "whole property. Fixed to require the settings module (same pattern as the earlier "
     "PATCH /auth/property/ fix). See Bug & observations sheet."),
    ("Method",
     "In-process via DRF's APIClient + force_authenticate (no live server, no auth-throttle limits, no "
     "test-data collisions across repeated runs). RBAC-read's endpoint catalogue was discovered by "
     "introspecting the live URLconf for every module-gated GET-list view, not hand-typed — so it "
     "reflects the actual routing table, and each (role, endpoint) pair appears exactly once. RBAC "
     "expectations are derived live from apps.accounts.rbac.can_access() (the same DB-aware check, "
     "including RoleConfig overrides, that the API itself enforces), never a hardcoded table."),
    ("De-duplication",
     "The original 15 Jul run had ~165 accidental duplicate RBAC-read rows (11 endpoints tested twice "
     "per role) from its own generator, plus only covered 15 of the app's 17 roles (missing Hotel "
     "Manager and HR Manager). This run's catalogue is built from a de-duplicated endpoint list — every "
     "case appears exactly once — and covers all 17 roles."),
    ("Regression safety", "Full backend test suite (335 tests, every app) green; full re-run of the "
     "100-case E2E suite green; this 1000-case suite itself reproduces identically pass/pass across "
     "repeated runs (1421/1421 twice in a row)."),
    ("Environment note", "Dev database; QA objects use QA-prefixed names; currency/entitlement/property "
     "fields are always restored to their original values after each round-trip check."),
]
r = 5
for label, val in rows:
    ws.cell(row=r, column=2, value=label).font = Font(bold=True, size=10, color=NAVY)
    c = ws.cell(row=r, column=3, value=val)
    c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
ws.cell(row=r, column=2, value="Cases by family").font = Font(bold=True, size=10, color=NAVY)
r += 1
ws.cell(row=r, column=2, value="Family").font = Font(bold=True, size=10)
ws.cell(row=r, column=3, value="What it covers").font = Font(bold=True, size=10)
ws.cell(row=r, column=4, value="Cases").font = Font(bold=True, size=10)
ws.cell(row=r, column=5, value="Pass").font = Font(bold=True, size=10)
r += 1
by_family = {}
for rec_ in RESULTS:
    fam = rec_["family"]
    by_family.setdefault(fam, [0, 0])
    by_family[fam][0] += 1
    if rec_["status"] == "PASS":
        by_family[fam][1] += 1
for fam in FAMILY_ORDER:
    if fam not in by_family:
        continue
    total, passed = by_family[fam]
    ws.cell(row=r, column=2, value=fam)
    desc_cell = ws.cell(row=r, column=3, value=FAMILY_DESC.get(fam, ""))
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4, value=total)
    ws.cell(row=r, column=5, value=passed)
    r += 1

# ---------------- All cases ----------------
ws2 = wb.create_sheet("All cases")
headers = ["Case ID", "Family", "Test Case", "Expected", "Actual", "Status"]
widths = [10, 14, 60, 26, 60, 8]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

for i, rec_ in enumerate(RESULTS, start=2):
    ws2.cell(row=i, column=1, value=rec_["id"])
    ws2.cell(row=i, column=2, value=rec_["family"])
    ws2.cell(row=i, column=3, value=rec_["desc"])
    ws2.cell(row=i, column=4, value=str(rec_["expected"]))
    actual_cell = ws2.cell(row=i, column=5, value=str(rec_["actual"]))
    actual_cell.alignment = Alignment(wrap_text=True, vertical="top")
    status_cell = ws2.cell(row=i, column=6, value=rec_["status"])
    if rec_["status"] == "PASS":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
        status_cell.font = Font(color=GREEN_FONT, bold=True)
    else:
        status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        status_cell.font = Font(color="B91C1C", bold=True)

# ---------------- Bug & observations ----------------
ws3 = wb.create_sheet("Bug & observations")
headers3 = ["ID", "Severity", "What", "Detail", "Fix / status"]
widths3 = [12, 16, 40, 70, 50]
for i, (h, w) in enumerate(zip(headers3, widths3), start=1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

bug_rows = [
    ("CX-BUG-02", "High (security)", "Entitlements editable by any logged-in user",
     "PATCH /auth/entitlements/ was permission-classed IsAuthenticated only, despite its own docstring "
     "claiming \"MD/GM only enforced client+server\" — no role/module check actually ran. A "
     "Housekeeping/cashier/captain login could disable HMS, Restaurant, Banquets or RMS for the whole "
     "property. Surfaced by this run's Entitlement family (non-settings role got 200 instead of 403).",
     "FIXED — PATCH now requires the settings module (IsAuthenticated + ModulePermission), same pattern "
     "as the earlier CX-BUG-01 fix for PATCH /auth/property/. backend/apps/accounts/views.py"),
    ("CX-OBS-03", "Low", "Re-checkout on a settled folio is a silent no-op",
     "POST /folios/{id}/checkout/ on an already-settled folio doesn't error — it returns 200 with the "
     "same invoice/settlement (balance is already 0, so check_out() has nothing to do). Safe (no double "
     "charge, no duplicate invoice) but gives a front-desk agent no feedback that the action was a no-op.",
     "OBSERVATION — consider a distinct message (\"already checked out\") for clarity; not a data-"
     "integrity issue."),
    ("CX-OBS-04", "Low", "kds/online modules still absent from ALL_MODULES",
     "Carried forward from the 15 Jul run: these modules gate real endpoints via the role allow-lists "
     "but aren't in ALL_MODULES, so the Role-Matrix admin screen has no toggle for them. Access control "
     "itself still works correctly.",
     "OBSERVATION — add to ALL_MODULES if per-role editing of KDS/Online in the Role Matrix UI is wanted."),
]
for i, row in enumerate(bug_rows, start=2):
    for j, val in enumerate(row, start=1):
        c = ws3.cell(row=i, column=j, value=val)
        if j in (4, 5):
            c.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}: {n_pass}/{n_total} pass, {len(bug_rows)} bugs/observations")
